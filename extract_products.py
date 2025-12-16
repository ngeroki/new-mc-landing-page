from openpyxl import load_workbook
import json

# Load the workbook
wb = load_workbook('public/website.xlsx')
ws = wb.active

# Define headers based on row 2
headers = ['No', 'Nama Produk', 'MOQ', 'Ukuran', 'Harga Flat', 'Harga High', 
           'Harga Jual Satuan Flat', 'Harga Jual Satuan High', 'Harga Grosir', 
           'Gender', 'Catatan Ukuran', 'Keterangan']

# Extract products starting from row 5
products = []
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    if row[0]:  # Only include rows with a number
        product = {headers[i]: row[i] for i in range(len(headers))}
        products.append(product)

print(f'Total products found: {len(products)}')
print('\nFirst 10 products:')
for i, p in enumerate(products[:10], 1):
    print(f"{i}. {p['Nama Produk']} - {p['Gender']} - Size: {p['Ukuran']} - Price: Rp {p['Harga Grosir']:,}")

# Save to JSON for easier processing
with open('products_data.json', 'w', encoding='utf-8') as f:
    json.dump(products, f, indent=2, ensure_ascii=False)

print(f'\nData saved to products_data.json')
